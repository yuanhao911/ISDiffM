
import os
import numpy as np
import torch
from openpyxl import load_workbook
from openpyxl import Workbook
import csv
import pickle

length_signal = 100
batch_size = 1024     # 每次训练的batch大小
nyquist_size = 51   # 阻抗谱离散点长度 51
nyquist_size_one_stage = 8   # 阻抗谱离散点长度
nyquist_seq = nyquist_size + 1   # 阻抗谱离散点序列长度
feature_size = 9           # 输入特征个数8
nyquist_feature_size = 2    # 输出特征个数 2
nyquist_feature_size_one_stage = 1   # 输出特征个数
epoch = 1000
simple_interval = 10 # 10
num_timesteps = 1000
num_proposals = 51


# tools functions
def import_dataset_x(path):
    res = []
    for filename in os.listdir(path):
        filepath = os.path.join(path, filename)
        wb = load_workbook(filename=filepath, read_only=True)
        ws = wb.active
        once = []
        for row in ws.iter_rows(values_only=True):
            float_list = [float(element) for element in row]
            once.append(float_list)
        res.append(once)
    res_tensor = torch.tensor(res).transpose(2, 1)
    return res_tensor


def import_dataset_y(path):
    res = []
    for filename in os.listdir(path):
        filepath = os.path.join(path, filename)
        with open(filepath, 'r') as file:
            reader = csv.reader(file)
            col0 = []
            col1 = []
            col2 = []
            for row in reader:
                col0.append(float(row[0]))
                col1.append(float(row[1])*1000)
                col2.append(float(row[2])*1000)
            once = [col1, col2]
        res.append(once)
    res_tensor = torch.tensor(res).transpose(2, 1)
    return res_tensor


def make_data(X, Y, Y_mean, Y_std):
    a, b, c = Y.shape[0], Y.shape[1], Y.shape[2]
    start = torch.tensor(start_symbol).expand(a, -1, -1)
    start = to_normalization(start, Y_mean, Y_std)
    end = torch.tensor(end_symbol).expand(a, -1, -1)
    end = to_normalization(end, Y_mean, Y_std)
    Y_dec_input = torch.cat((start, Y), dim=1)
    Y_dec_output = torch.cat((Y, end), dim=1)
    X_enc_input = X
    return X_enc_input, Y_dec_input, Y_dec_output


def simple(X):
    X_simple = X[:, ::simple_interval, :]
    predict = length_signal
    X_simple = X_simple[:, 180-predict:180, :]
    return X_simple


def normalization(input):
    input_mean = torch.mean(input, dim=(0, 1))
    input_std = torch.std(input, dim=(0, 1))
    input_normalized = (input - input_mean) / input_std
    return input_normalized, input_mean, input_std


def to_normalization(input, mean, std):
    output = (input - mean) / std
    return output


def denormalization(input_normalized, input_mean, input_std):
    output = input_normalized * input_std + input_mean
    return output


def save_variable(v, filename):
    f = open(filename, 'wb')
    pickle.dump(v, f)
    f.close()
    return filename


def load_variavle(filename):
    f = open(filename, 'rb')
    r = pickle.load(f)
    f.close()
    return r


def add_noise_1(x, snr_db):
    x_array = x.numpy()
    P_signal = np.mean(x_array ** 2, axis=1)
    P_noise = P_signal / 10 ** (snr_db / 10.0)
    d = np.random.randn(x.shape[1])
    noise = P_noise[:, :, np.newaxis] * d[np.newaxis, :]
    noise = noise.transpose(0, 2, 1)
    x_array_noise = x_array + noise
    x_tesnsor_noise = torch.from_numpy(x_array_noise).float()
    return x_tesnsor_noise


def remove_axis(x, num):
    x_remove = x.clone()
    x_remove[:, :, num].fill_(0)
    return x_remove
